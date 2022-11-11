import openpyxl
#open excel file
wb = openpyxl.load_workbook("excel.xlsx")
ws = wb['Sheet1']#wb.active
rows = ws.iter_rows(min_row=1, max_row=1)
d = []
K = []

#headers
for row in rows:
    l = []
    for a in row:
        l.append(a.value)
    d.append(l)

# row 3 to all
rows = ws.iter_rows(min_row=3, max_row=ws.max_row)
for row in rows:
    l = []
    for a in row:
        s=a.value
        if s== None:
            s=""
        l.append(s)
    lis =[]
    for i in range(len(d[0])):
        if l[9] == "":
            l[9] =l[2]
            # if l[2]=="":
            #     l[9]=l[2]
           #  if l[2] == None:
           #      l[9] = ""


        row_list = "<ROW RECORDID=00 MODID=000000>"+"<" + str(d[0][5]) + ">" + str(l[5]) + "</" + str(d[0][5]) + ">"+"<journalfullTitle>"+"<DATA>"+str(l[5])+"</DATA>"+"</JournalFullTitle>"+"<"+str(d[0][6])+">"+str(l[6])+"</"+str(d[0][6])+">"+"<"+str(d[0][7])+">"+str(l[7])+"</"+str(d[0][7])+">"+"<"+str(d[0][3])+">"+str(l[3])+"</"+str(d[0][3])+">"+"<"+str(d[0][21])+">"+str(l[21])+"</"+str(d[0][21])+">"+"<"+str(d[0][22])+">"+str(l[22])+"</"+str(d[0][22])+">"+"<"+str(d[0][11])+">"+str(l[11])+"</"+str(d[0][11])+">"+"<"+str(d[0][18])+">"+str(l[18])+"</"+str(d[0][18])+">"+"<"+str(d[0][10])+">"+str(l[10])+"</"+str(d[0][10])+">"+"<"+str(d[0][24])+">"+str(l[24])+"</"+str(d[0][24])+">"+"<"+str(d[0][9])+">"+str(l[9])+"</"+str(d[0][9])+">"+"</ROW>"
        lis.append(row_list)
    L="".join(lis)
    K.append(L)
F="<articles>"+''.join(K)+"</articles>"
print(F)
with open("Final.xml", "w",encoding='UTF-8') as myfile:
    myfile.write(F)